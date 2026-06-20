#!/usr/bin/env python3
"""Pull team roster from Google Sheets and provide fuzzy name matching."""

import io
import json
import os
from difflib import SequenceMatcher

import openpyxl
from google.oauth2.credentials import Credentials
from google.auth.transport.requests import Request
from googleapiclient.discovery import build

from config import GOOGLE_CLIENT_ID, GOOGLE_CLIENT_SECRET

TOKEN_PATH = os.path.join(os.path.dirname(__file__), 'token.json')
SHEET_FILE_ID = os.getenv('TEAM_SHEET_ID')

# Map each project name (from config.py PROJECTS) to the exact tab name in your
# Google Sheet roster. Set to None if a project has no dedicated roster tab.
PROJECT_TAB_MAP = {
    'Example Project Alpha': 'Example Project Alpha',
    'Example Project Beta': 'Example Project Beta',
}


def get_credentials():
    with open(TOKEN_PATH) as f:
        td = json.load(f)
    creds = Credentials(
        token=td['token'],
        refresh_token=td['refresh_token'],
        token_uri=td['token_uri'],
        client_id=GOOGLE_CLIENT_ID,
        client_secret=GOOGLE_CLIENT_SECRET,
    )
    if creds.expired and creds.refresh_token:
        creds.refresh(Request())
        td['token'] = creds.token
        with open(TOKEN_PATH, 'w') as f:
            json.dump(td, f, indent=2)
    return creds


def download_workbook():
    creds = get_credentials()
    drive = build('drive', 'v3', credentials=creds)
    content = drive.files().get_media(fileId=SHEET_FILE_ID).execute()
    return openpyxl.load_workbook(io.BytesIO(content), data_only=True)


def get_master_roster(wb=None):
    """Get all team members from the 'Team Members' tab."""
    if wb is None:
        wb = download_workbook()

    ws = wb['Team Members']
    roster = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        full_name = row[0]
        email = row[3] if len(row) > 3 else None
        if full_name and email:
            roster.append({
                'full_name': str(full_name).strip(),
                'email': str(email).strip(),
                'role': str(row[1]).strip() if row[1] else '',
                'team': str(row[2]).strip() if row[2] else '',
            })
    return roster


def get_project_members(project_name, wb=None):
    """Get members for a specific project from its dedicated tab."""
    if wb is None:
        wb = download_workbook()

    tab_name = PROJECT_TAB_MAP.get(project_name)
    if not tab_name or tab_name not in wb.sheetnames:
        return []

    ws = wb[tab_name]
    headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows(min_row=1, max_row=1))]

    name_col = None
    email_col = None
    for i, h in enumerate(headers):
        if 'full name' in h or h == 'full name':
            name_col = i
        elif h == 'first name':
            name_col = i
        if 'e-mail' in h or 'email' in h:
            email_col = i

    if name_col is None:
        return []

    last_name_col = None
    for i, h in enumerate(headers):
        if 'last name' in h:
            last_name_col = i

    members = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if last_name_col is not None:
            first = str(row[name_col]).strip() if row[name_col] else ''
            last = str(row[last_name_col]).strip() if row[last_name_col] else ''
            full_name = f'{first} {last}'.strip()
        else:
            full_name = str(row[name_col]).strip() if row[name_col] else ''

        email = str(row[email_col]).strip() if email_col is not None and row[email_col] else ''

        if full_name:
            members.append({
                'full_name': full_name,
                'email': email,
            })

    return members


def fuzzy_match_name(transcript_name, roster, threshold=0.6):
    """Match a transcript speaker name to the closest roster entry."""
    transcript_name = transcript_name.strip()
    best_match = None
    best_score = 0

    for member in roster:
        real_name = member['full_name']

        score = SequenceMatcher(None, transcript_name.lower(), real_name.lower()).ratio()

        first_name = real_name.split()[0] if real_name.split() else ''
        first_score = SequenceMatcher(None, transcript_name.lower(), first_name.lower()).ratio()

        final_score = max(score, first_score)

        if final_score > best_score:
            best_score = final_score
            best_match = member

    if best_score >= threshold:
        return best_match
    return None


def correct_transcript_names(transcript_text, project_name):
    """Replace misspelled names in transcript with correct roster names."""
    wb = download_workbook()
    project_members = get_project_members(project_name, wb)
    master_roster = get_master_roster(wb)
    all_names = {m['full_name']: m for m in master_roster}
    for pm in project_members:
        if pm['full_name'] not in all_names:
            all_names[pm['full_name']] = pm
    roster = list(all_names.values())

    import re
    speaker_pattern = re.compile(r'^([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)\s*:', re.MULTILINE)

    corrections = {}

    def replace_speaker(match):
        name = match.group(1)
        if name in corrections:
            return f'{corrections[name]}:'

        matched = fuzzy_match_name(name, roster)
        if matched and matched['full_name'] != name:
            corrections[name] = matched['full_name']
            return f'{matched["full_name"]}:'
        return match.group(0)

    corrected = speaker_pattern.sub(replace_speaker, transcript_text)

    if corrections:
        print(f'Name corrections applied:')
        for wrong, right in corrections.items():
            print(f'  "{wrong}" → "{right}"')

    return corrected, corrections


if __name__ == '__main__':
    wb = download_workbook()
    roster = get_master_roster(wb)
    print(f'Master roster: {len(roster)} members\n')
    for m in roster[:5]:
        print(f'  {m["full_name"]} ({m["email"]})')
    print(f'  ... and {len(roster) - 5} more\n')

    for project_name, tab in PROJECT_TAB_MAP.items():
        members = get_project_members(project_name, wb)
        print(f'{project_name}: {len(members)} members')
